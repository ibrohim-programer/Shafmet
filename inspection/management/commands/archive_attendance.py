import os
import zoneinfo
from datetime import datetime
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.conf import settings
from inspection.models import Attendance, DailyAttendance
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class Command(BaseCommand):
    help = "Uzbekistan vaqti bilan kunlik davomatlarni Excel formatida arxivlab, so'ngra bazadan o'chiradi."

    def add_arguments(self, parser):
        parser.add_argument(
            "--date",
            type=str,
            help="Arxivlash uchun sana (YYYY-MM-DD formatida). Kiritilmasa, bugungi sana olinadi.",
        )
        parser.add_argument(
            "--output-dir",
            type=str,
            default="/home/ubuntu/Documents",
            help="Arxiv fayli saqlanadigan papka yo'li.",
        )

    def handle(self, *args, **options):
        # O'zbekiston vaqt mintaqasi (Asia/Tashkent - UTC+5)
        tashkent_tz = zoneinfo.ZoneInfo("Asia/Tashkent")
        now_tashkent = timezone.localtime(timezone.now(), tashkent_tz)

        # Sanani aniqlash
        date_str = options["date"]
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                self.stderr.write(
                    self.style.ERROR(
                        f"Xato sana formati: '{date_str}'. YYYY-MM-DD formatidan foydalaning."
                    )
                )
                return
        else:
            target_date = now_tashkent.date()

        self.stdout.write(
            f"Keltirilgan sana uchun davomatlar yig'ilmoqda: {target_date}"
        )

        # 1. Attendance (Davomat) yozuvlarini olish
        attendance_qs = Attendance.objects.filter(date=target_date)

        # 2. DailyAttendance (Kunlik davomat) yozuvlarini olish
        daily_qs = DailyAttendance.objects.filter(date=target_date)

        # Agar hech qanday davomat yozuvi bo'lmasa, fayl yaratib o'tirmaymiz
        if not attendance_qs.exists() and not daily_qs.exists():
            self.stdout.write(
                self.style.WARNING(
                    f"{target_date} sanasida hech qanday davomat ma'lumotlari topilmadi. Amallar bajarilmadi."
                )
            )
            return

        # Excel ishchi kitobini (Workbook) yaratish
        wb = Workbook()

        # 1-sahifa: Batafsil Davomat
        ws1 = wb.active
        ws1.title = "Batafsil Davomat"

        # 2-sahifa: Kunlik Davomat
        ws2 = wb.create_sheet(title="Kunlik Davomat")

        # Stillar
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # 1-sahifa ustun nomlari
        headers1 = [
            "ID",
            "Xodim (Ism Familiya)",
            "Telefon",
            "Sana",
            "Kirish vaqti",
            "Kirish muvaffaqiyatli",
            "Chiqish vaqti",
            "Chiqish muvaffaqiyatli",
            "Kechikdimi",
        ]
        ws1.append(headers1)

        # 1-sahifa ma'lumotlarini to'ldirish
        for att in attendance_qs.select_related("worker"):
            ws1.append(
                [
                    att.id,
                    att.worker.full_name,
                    att.worker.phone,
                    att.date.isoformat() if att.date else "",
                    (
                        timezone.localtime(att.check_in_time, tashkent_tz).strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        if att.check_in_time
                        else "-"
                    ),
                    "Ha" if att.check_in_success else "Yo'q",
                    (
                        timezone.localtime(att.check_out_time, tashkent_tz).strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        if att.check_out_time
                        else "-"
                    ),
                    (
                        "Ha"
                        if att.check_out_success
                        else (
                            "Yo'q" if att.check_out_success is False else "-"
                        )
                    ),
                    "Ha" if att.is_late else "Yo'q",
                ]
            )

        # 2-sahifa ustun nomlari
        headers2 = [
            "ID",
            "Foydalanuvchi (Ism Familiya)",
            "Telefon",
            "Sana",
            "Kirish vaqti",
            "Chiqish vaqti",
            "Kechikdimi",
        ]
        ws2.append(headers2)

        # 2-sahifa ma'lumotlarini to'ldirish
        for d_att in daily_qs.select_related("user"):
            ws2.append(
                [
                    d_att.id,
                    d_att.user.full_name,
                    d_att.user.phone,
                    d_att.date.isoformat() if d_att.date else "",
                    (
                        timezone.localtime(d_att.check_in_time, tashkent_tz).strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        if d_att.check_in_time
                        else "-"
                    ),
                    (
                        timezone.localtime(d_att.check_out_time, tashkent_tz).strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        if d_att.check_out_time
                        else "-"
                    ),
                    "Ha" if d_att.is_late else "Yo'q",
                ]
            )

        # Ikkala sahifaga stillar va kengliklarni moslash
        for ws in [ws1, ws2]:
            # Sarlavhalarga stil berish
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # Kataklarni tekislash va ustun kengliklarini avtomatik hisoblash
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    # Ma'lumot qatorlari uchun
                    if cell.row > 1:
                        if (
                            isinstance(cell.value, str)
                            and cell.value not in ["Ha", "Yo'q", "-"]
                        ):
                            cell.alignment = left_align
                        else:
                            cell.alignment = center_align
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col[0].column_letter].width = max(
                    max_len + 3, 12
                )

        filename = f"{target_date.isoformat()}.xlsx"

        # 1. Serverdagi arxiv papkasi (Loyiha ichida) - barcha adminlar yuklab olishi uchun
        primary_dir = os.path.join(settings.BASE_DIR, "archives")
        if not os.path.exists(primary_dir):
            try:
                os.makedirs(primary_dir, exist_ok=True)
            except Exception as e:
                self.stderr.write(
                    self.style.ERROR(
                        f"Server arxiv papkasini yaratib bo'lmadi ({primary_dir}): {e}"
                    )
                )
                return

        primary_filepath = os.path.join(primary_dir, filename)

        # Excel faylini serverga saqlash
        try:
            wb.save(primary_filepath)
            self.stdout.write(
                self.style.SUCCESS(
                    f"Muvaffaqiyatli saqlandi (Server arxivi): {primary_filepath}"
                )
            )
        except Exception as e:
            self.stderr.write(
                self.style.ERROR(
                    f"Server arxiviga yozishda xatolik yuz berdi ({primary_filepath}): {e}"
                )
            )
            return

        # 2. Laptopdagi Documents papkasi (agar u mavjud bo'lsa va yozish ruxsati bo'lsa)
        secondary_dir = options["output_dir"]
        if os.path.exists(secondary_dir) and os.access(secondary_dir, os.W_OK):
            secondary_filepath = os.path.join(secondary_dir, filename)
            if os.path.abspath(primary_filepath) != os.path.abspath(secondary_filepath):
                try:
                    wb.save(secondary_filepath)
                    self.stdout.write(
                        self.style.SUCCESS(
                            f"Muvaffaqiyatli nusxalandi (Laptop hujjatlari): {secondary_filepath}"
                        )
                    )
                except Exception as e:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Laptop hujjatlariga nusxalashda xatolik yuz berdi ({secondary_filepath}): {e}"
                        )
                    )

        # Bazadan o'chirish
        try:
            att_deleted, _ = attendance_qs.delete()
            daily_deleted, _ = daily_qs.delete()
            self.stdout.write(
                self.style.SUCCESS(
                    f"Baza tozalandi: {att_deleted} ta Attendance va {daily_deleted} ta DailyAttendance o'chirildi."
                )
            )
        except Exception as e:
            self.stderr.write(
                self.style.ERROR(
                    f"Bazadan davomatlarni o'chirishda xatolik: {e}"
                )
            )
